from bs4 import BeautifulSoup
import urllib
import urllib.request
import csv
from openpyxl import Workbook, load_workbook

def innerHTML(element):
    return element.decode_contents(formatter="html")

def get_name(body):
    return body.find('span', {'class':'jcn'}).a.string

def get_phone_number(body):
    try:
        phone_element = body.find('a', {'href': lambda x: x and x.startswith('tel:')})
        if phone_element:
            return phone_element.text.strip()
        return ''
    except AttributeError:
        return ''

def get_rating(body):
    rating = 0.0
    text = body.find('span', {'class':'star_m'})
    if text is not None:
        for item in text:
            rating += float(item['class'][0][1:])/10

    return rating

def get_rating_count(body):
    text = body.find('span', {'class':'rt_count'}).string

    # Get only digits
    rating_count =''.join(i for i in text if i.isdigit())
    return rating_count

def get_address(body):
    return body.find('span', {'class':'mrehover'}).text.strip()

def get_location(body):
    text = body.find('a', {'class':'rsmap'})
    if text == None:
        return
    text_list = text['onclick'].split(",")
    
    latitude = text_list[3].strip().replace("'", "")
    longitude = text_list[4].strip().replace("'", "")
    
    return latitude + ", " + longitude

def write_to_excel(data, filename):
    """
    Writes the given data to an Excel file. If the file exists, appends data to it.

    :param data: List of dictionaries containing the data.
    :param filename: Name of the Excel file to save.
    """
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Scraped Data"
        # Write header if new file
        headers = ['Name', 'Phone', 'Rating', 'Rating Count', 'Address', 'Location']
        sheet.append(headers)

    # Write data rows
    for row in data:
        sheet.append([row.get(header, '') for header in ['Name', 'Phone', 'Rating', 'Rating Count', 'Address', 'Location']])

    # Save the workbook
    workbook.save(filename)

page_number = 1
service_count = 1
scraped_data = []

fields = ['Name', 'Phone','Rating', 'Rating Count', 'Address', 'Location']
out_file = open('arunachal_pradesh.csv', 'w', newline='')
csvwriter = csv.DictWriter(out_file, delimiter=',', fieldnames=fields)

# Write fields first
csvwriter.writeheader()

urls = [
    "https://www.justdial.com/Chengalpattu/Schools-For-Deaf-Dumb/nct-11603364-%s"
]
city = "gurugram"

# Loop through all URLs and scrape data
for url_template in urls:
    page_number = 1
    while True:
        url = url_template % (page_number)
        req = urllib.request.Request(url, headers={'User-Agent' : "Magic Browser"}) 
        page = urllib.request.urlopen(req)

        soup = BeautifulSoup(page.read(), "html.parser")
        services = soup.find_all('li', {'class': 'cntanr'})

        # Iterate through the results in the page
        for service_html in services:
            # Parse HTML to fetch data
            dict_service = {}
            name = get_name(service_html)
            phone = get_phone_number(service_html)
            rating = get_rating(service_html)
            count = get_rating_count(service_html)
            address = get_address(service_html)
            location = get_location(service_html)
            
            if name:
                dict_service['Name'] = name
            if phone:
                dict_service['Phone'] = phone
            if rating:
                dict_service['Rating'] = rating
            if count:
                dict_service['Rating Count'] = count
            if address:
                dict_service['Address'] = address
            if location:
                dict_service['Location'] = location

            # Write row to CSV
            csvwriter.writerow(dict_service)
            scraped_data.append(dict_service)

            print("#" + str(service_count) + " ", dict_service)
            service_count += 1

        page_number += 1
        # Add condition to break loop when no more pages are available (optional improvement)
        break

# Write all collected data to Excel after scraping all URLs
write_to_excel(scraped_data, "DeafDumbSchool.xlsx")